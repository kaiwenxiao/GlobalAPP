import os
import unittest

import openpyxl
import parameterized
import paramunittest

import readConfig
from Common import Log
from Common import common, configHttp

RegisterInfo_xls = common.get_xls("testRegister.xlsx", "cardCheck")
localReadConfig = readConfig.ReadConfig()
configHttp = configHttp.ConfigHttp()
info = {}
proDir = readConfig.proDir
common.get_register_data_from_other_sheet("testRegister.xlsx")


# parameterized 需要第一行用例信息作为参数名
@paramunittest.parametrized(*RegisterInfo_xls)
class TestCardCheck(unittest.TestCase):

    def setParameters(self, case_name, method, membershipNo, pwd, activationNote, dragoncardNote, note):
        self.case_name = str(case_name)
        self.method = str(method)
        self.membershipNo = str(membershipNo)
        self.pwd = str(pwd)
        self.activationNote = str(activationNote)
        self.dragoncardNote = str(dragoncardNote)
        self.note = str(note)
        self.return_json = None
        self.info = None

    def description(self):
        self.case_name

    def setUp(self):
        self.log = Log.MyLog.get_log()
        self.logger = self.log.get_logger()
        print(self.case_name + "测试开始前准备")

    def testCardCheck(self):
        self._testMethodDoc = self.case_name
        self.url = common.get_url_from_xml("cardCheck")
        configHttp.set_url(self.url)
        print("1.设置url")

        header = {
            'unEncryptToken': 'true',
            'Content-Type': 'application/x-www-form-urlencoded',
        }
        configHttp.set_headers(header)
        print("2.设置header")

        data = {
            'membershipNo': self.membershipNo,
            'pwd': self.pwd
        }
        print(type(data))
        configHttp.set_data(data)
        print("3.设置发送参数")
        self.return_json = configHttp.post()
        method = str(self.return_json.request)[
                 int(str(self.return_json.request).find('[')) + 1:int(str(self.return_json.request).find(']'))]
        print("4.发送请求\n\t\t请求方法:" + method)

        self.checkResult()
        print("5.检查结果")

    def tearDown(self):
        self.log.build_case_line(self.case_name, self.note)

    def checkResult(self):
        self.info = self.return_json.json()
        print(self.info)
        # common.show_return_msg()
        if self.dragoncardNote == None:
            self.assertEqual(self.info['dragoncardNote'], self.dragoncardNote)
        if self.activationNote == None:
            self.assertEqual(self.info['activationNote'], self.activationNote)

        print(self.case_name)
        if self.case_name == "avaliable_membership":
            self.assertIsNotNone(self.info['uuid'])
            xlsPath = os.path.join(proDir, "testFile", "case", "testRegister.xlsx")
            file = openpyxl.load_workbook(xlsPath)
            data_sheet = file.worksheets[1]
            dynamic_no_used_row = str(int(data_sheet.cell(row=2, column=4).value) + 1)
            data_sheet.cell(row=2, column=4).value = dynamic_no_used_row
            file.save(str(xlsPath))


if __name__ == "__main__":
    ProductInfo1 = TestCardCheck()
