import json
import os
from xml.etree import ElementTree

from xlrd import open_workbook
import openpyxl
import readConfig

proDir = readConfig.proDir


# 读取Excel里workSheet的测试用例参数
def get_xls(xls_name, sheet_name):
    cls = []
    xlsPath = os.path.join(proDir, "testFile", "case", xls_name)
    file = open_workbook(xlsPath)
    sheet = file.sheet_by_name(sheet_name)
    nrows = sheet.nrows
    for i in range(nrows):
        # 除第一行的第一格为测试用例名字
        if sheet.row_values(i)[0] != u'case_name':
            cls.append(sheet.row_values(i))
    return cls


def get_url_from_xml(name):
    url_list = []
    url_path = os.path.join(proDir, 'testFile', 'interfaceURL.xml')
    tree = ElementTree.parse(url_path)
    for u in tree.findall('url'):
        url_name = u.get('name')
        if url_name == name:
            for c in u.getchildren():
                url_list.append(c.text)
    url = '/VirtualCard-en/' + '/'.join(url_list)
    return url


def get_register_data_from_other_sheet(xls_name):
    # xlsPath = os.path.join(proDir, "testFile", "case", xls_name)
    # file = open_workbook(xlsPath)
    # orginal_sheet = file.get_sheet(0)
    # data_sheet = file.get_sheet(1)
    # data_sheet.write(4, 2, data_sheet.cell(1, 1).value)
    # data_sheet.write(4, 3, data_sheet.cell(1, 2).value)
    # # orginal_sheet.row_values(4)[2] = data_sheet.cell(1, 1).value
    # # orginal_sheet.row_values(4)[3] = data_sheet.cell(1, 2).value
    # print("fuc")
    # print(orginal_sheet.row_values(4)[2])
    # print(orginal_sheet.row_values(4)[3])

    xlsPath = os.path.join(proDir, "testFile", "case", xls_name)
    file = openpyxl.load_workbook(xlsPath)
    data_sheet = file.worksheets[1]
    orginal_sheet = file.worksheets[0]

    print(data_sheet.max_row)
    no_used_row = int(data_sheet.cell(row=2, column=4).value)
    no_used_membership = data_sheet.cell(row=no_used_row, column=2)
    no_used_activecode = data_sheet.cell(row=no_used_row, column=3)
    orginal_sheet.cell(row=5, column=3).value = no_used_membership.value
    orginal_sheet.cell(row=5, column=4).value = no_used_activecode.value

    file.save(str(xlsPath))


def show_return_msg(response):
    url = response.url
    msg = response.text
    print("\n请求地址:" + url)
    # 可以显示中文
    print("\n请求返回值：" + '\n' + json.dumps(json.loads(msg), ensure_ascii=False, sort_keys=True, indent=4))


if __name__ == "__main__":
    get_register_data_from_other_sheet("testRegister.xlsx")
